from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.http import HttpResponse, JsonResponse
from io import BytesIO
import re
from collections import defaultdict
from django.db.models import Q
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from .models import ColorProfile, Student, Subject, Grade, Vedomost, UserProfile, TeachingAssignment, TeacherAssignmentFile
from .forms import UploadFileForm, CreateTeacherForm, ReportGenerationForm, EditUserForm, VedomostFilterForm
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import os
import hashlib
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction
import logging
logger = logging.getLogger(__name__)

# ---------------------- AUTH ----------------------
def user_login(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('home')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

@login_required
def user_logout(request):
    logout(request)
    return redirect('login')

def home_view(request):
    if request.user.is_authenticated:
        if request.user.userprofile.role == 'teacher':
            return redirect('upload')
        elif request.user.userprofile.role == 'deputy':
            return redirect('vedomosti_list')
    return render(request, 'home.html')

# ---------------------- TEACHER ----------------------
@login_required
def upload_file(request):
    if request.user.userprofile.role != 'teacher':
        messages.error(request, "Загрузка ведомостей доступна только преподавателям.")
        return redirect('home')

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            fs = FileSystemStorage(location='media/vedomosti/')
            filename = fs.save(file.name, file)
            file_path = fs.path(filename)

            try:
                parse_and_save(request.user, file_path)
                messages.success(request, "Файл успешно обработан.")
                return redirect('upload')
            except ValueError as e:
                if os.path.exists(file_path):
                    os.remove(file_path)
                messages.error(request, str(e))
        else:
            # вот здесь добавим общую ошибку (например, неверный формат)
            messages.error(request, "Ошибка загрузки: " + "; ".join(
                error for errors in form.errors.values() for error in errors
            ))
    else:
        form = UploadFileForm()

    return render(request, 'upload.html', {'form': form})


@login_required
def color_settings(request):
    if request.user.userprofile.role != 'teacher':
        messages.error(request, "Настройки доступны только преподавателям.")
        return redirect('home')

    profile, _ = ColorProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        profile.student_color = request.POST.get('student_color', profile.student_color)
        profile.subject_color = request.POST.get('subject_color', profile.subject_color)
        profile.grade_color = request.POST.get('grade_color', profile.grade_color)
        profile.group_color = request.POST.get('group_color', profile.group_color)
        profile.period_color = request.POST.get('period_color', profile.period_color)
        profile.save()
        return redirect('upload')

    return render(request, 'color_settings.html', {'color_profile': profile})

# ---------------------- PARSING ----------------------
def generate_vedomost_hash(file_path):
    with open(file_path, 'rb') as f:
        return hashlib.sha256(f.read()).hexdigest()

def get_cell_rgb(cell):
    color = cell.fill.start_color
    if color.type == 'rgb' and color.rgb:
        return color.rgb.upper()
    return None

def parse_and_save(user, file_path):
    profile = ColorProfile.objects.get(user=user)

    def hex_to_rgb(hex_color):
        return "FF" + hex_color.lstrip("#").upper()

    student_color = hex_to_rgb(profile.student_color)
    subject_color = hex_to_rgb(profile.subject_color)
    grade_color = hex_to_rgb(profile.grade_color)
    group_color = hex_to_rgb(profile.group_color)
    period_color = hex_to_rgb(profile.period_color)

    wb = load_workbook(file_path)
    results = []

    with transaction.atomic():
        for sheet in wb.worksheets:
            relative_path = os.path.relpath(file_path, settings.MEDIA_ROOT)
            vedomost = Vedomost(
                title=f"{os.path.basename(file_path)} — {sheet.title}",
                file=relative_path,
                uploaded_by=user
            )

            subjects_by_col = {}
            students_by_row = {}
            found_grades = []

            for row in sheet.iter_rows():
                for cell in row:
                    value = str(cell.value).strip() if cell.value else None
                    if not value:
                        continue

                    rgb = get_cell_rgb(cell)
                    if rgb is None or rgb == "00000000":
                        continue

                    if rgb == subject_color:
                        subjects_by_col[cell.column] = value
                    elif rgb == student_color:
                        students_by_row[cell.row] = value
                    elif rgb == grade_color:
                        found_grades.append((cell.row, cell.column, value))
                    elif rgb == group_color:
                        vedomost.group_name = value
                    elif rgb == period_color:
                        match = re.search(r'За\s+(\d+)-й\s+семестр\s+(\d{4}-\d{4})\s+учебного\s+года', value)
                        if not match:
                            raise ValueError(f"Лист '{sheet.title}' Неверный формат периода: '{value}'")
                        vedomost.semester = match.group(1)
                        vedomost.academic_year = match.group(2)

            if not vedomost.group_name or not vedomost.academic_year or not vedomost.semester:
                raise ValueError(f"Лист '{sheet.title}' Не удалось распознать группу, семестр или учебный год (проверьте цвет).")

            if not students_by_row:
                raise ValueError(f"Лист '{sheet.title}' Не найдено ни одного студента (проверьте цвет ФИО).")

            if not subjects_by_col:
                raise ValueError(f"Лист '{sheet.title}' Не найдено ни одного предмета (проверьте цвет Предметов).")

            if not found_grades:
                raise ValueError(f"Лист '{sheet.title}' Не найдено ни одной оценки (проверьте цвет Оценок).")

            sheet_data = ''.join(f"{cell.value}" for row in sheet.iter_rows() for cell in row if cell.value)
            data_hash = hashlib.sha256(sheet_data.encode('utf-8')).hexdigest()

            if Grade.objects.filter(vedomost__data_hash=data_hash).exists():
                raise ValueError(f"Лист '{sheet.title}' Такая ведомость уже была загружена ранее.")

            if Vedomost.objects.filter(
                group_name=vedomost.group_name,
                semester=vedomost.semester,
                academic_year=vedomost.academic_year
            ).exists():
                raise ValueError(f"Лист '{sheet.title}' Ведомость для этой группы уже существует.")

            vedomost.data_hash = data_hash
            vedomost.save()

            students_cache = {}
            subjects_cache = {}

            for row_idx, col_idx, grade_value in found_grades:
                student_name = students_by_row.get(row_idx)
                subject_name = subjects_by_col.get(col_idx)
                if not student_name or not subject_name:
                    continue

                student = students_cache.get(student_name)
                if not student:
                    student, _ = Student.objects.get_or_create(full_name=student_name, group=vedomost.group_name)
                    students_cache[student_name] = student

                subject = subjects_cache.get(subject_name)
                if not subject:
                    try:
                        subject = Subject.objects.get(name=subject_name)
                    except Subject.DoesNotExist:
                        raise ValueError(f"Лист '{sheet.title}' Предмет '{subject_name}' не найден.")
                    subjects_cache[subject_name] = subject

                if not TeachingAssignment.objects.filter(subject=subject, group=vedomost.group_name).exists():
                    raise ValueError(
                        f"Лист '{sheet.title}' Дисциплина '{subject_name}' не назначена ни одному преподавателю для группы '{vedomost.group_name}'."
                    )

                Grade.objects.create(
                    vedomost=vedomost,
                    student=student,
                    subject=subject,
                    value=grade_value
                )

# ---------------------- DEPUTY ----------------------
@login_required
def vedomosti_list(request):
    if request.user.userprofile.role != 'deputy':
        messages.error(request, "Просмотр ведомостей доступен только завучу.")
        return redirect('home')

    vedomosti = Vedomost.objects.all().order_by('-upload_date')

    groups = vedomosti.values_list('group_name', flat=True).distinct()
    years = vedomosti.values_list('academic_year', flat=True).distinct()
    teachers = User.objects.filter(userprofile__role='teacher').select_related('userprofile')

    form = VedomostFilterForm(request.GET or None, vedomosti=vedomosti)

    if form.is_valid():
        group = form.cleaned_data.get('group')
        year = form.cleaned_data.get('academic_year')
        uploaded_by = form.cleaned_data['uploaded_by']

        if uploaded_by:
            vedomosti = vedomosti.filter(uploaded_by__id=uploaded_by)
        if group:
            vedomosti = vedomosti.filter(group_name=group)
        if year:
            vedomosti = vedomosti.filter(academic_year=year)

    vedomosti_data = []
    for ved in vedomosti:
        grades = Grade.objects.filter(vedomost=ved)
        student_ids = grades.values_list('student_id', flat=True).distinct()
        vedomosti_data.append({
            'ved': ved,
            'student_count': student_ids.count(),
            'grade_count': grades.count(),
        })

    return render(request, 'vedomosti.html', {
        'vedomosti_data': vedomosti_data,
        'form': form,
    })

@login_required
def grades_view(request, vedomost_id):
    vedomost = get_object_or_404(Vedomost, id=vedomost_id)
    grades = Grade.objects.filter(vedomost=vedomost).select_related('student', 'subject')

    grade_data = []
    for grade in grades:
        assignment = TeachingAssignment.objects.filter(
            subject=grade.subject,
            group=vedomost.group_name
        ).first()
        teacher = assignment.teacher.userprofile if assignment else None

        grade_data.append({
            'student': grade.student,
            'subject': grade.subject,
            'value': grade.value,
            'teacher': teacher,
        })

    return render(request, 'grades.html', {
        'vedomost': vedomost,
        'grades': grade_data,
    })

@login_required
def delete_vedomost(request, ved_id):
    if request.user.userprofile.role != 'deputy':
        messages.error(request, "Удаление доступно только завучу.")
        return redirect('home')

    vedomost = get_object_or_404(Vedomost, id=ved_id)
    vedomost.delete()
    return redirect('vedomosti_list')

@login_required
def create_teacher(request):
    if request.user.userprofile.role != 'deputy':
        messages.error(request, "Создание пользователей доступно только завучу.")
        return redirect('home')

    if request.method == 'POST':
        form = CreateTeacherForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Преподаватель успешно создан.")
            return redirect('create_teacher')
    else:
        form = CreateTeacherForm()
    return render(request, 'create_teacher.html', {'form': form})

@login_required
def color_settings_ta(request):
    if request.user.userprofile.role != 'deputy':
        messages.error(request, "Настройки доступны только завучу.")
        return redirect('home')

    profile, _ = ColorProfile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        profile.ta_subject_color = request.POST.get('ta_subject_color', profile.ta_subject_color)
        profile.ta_teacher_color = request.POST.get('ta_teacher_color', profile.ta_teacher_color)
        profile.ta_group_color = request.POST.get('ta_group_color', profile.ta_group_color)
        profile.save()
        messages.success(request, "Цвета для назначения преподавателей сохранены.")
        return redirect('color_settings_ta')

    return render(request, 'color_settings_ta.html', {'color_profile': profile})


@login_required
def upload_teacher_assignments(request):
    if request.user.userprofile.role != 'deputy':
        messages.error(request, "Доступ запрещен.")
        return redirect('home')

    form = UploadFileForm()
    context = {'form': form}

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            fs = FileSystemStorage(location=os.path.join(settings.MEDIA_ROOT, 'teacher_assignments'))
            filename = fs.save(file.name, file)
            full_path = fs.path(filename)

            try:
                # Сначала создаём объект файла
                assignment_file = TeacherAssignmentFile.objects.create(
                    file='teacher_assignments/' + filename,
                    uploaded_by=request.user
                )

                result_count = parse_teacher_assignments(request.user, full_path, assignment_file)

                if result_count == 0:
                    assignment_file.delete()
                    os.remove(full_path)
                    messages.error(request, "Файл не содержит новых назначений.")
                else:
                    messages.success(request, f"Назначения успешно загружены: {result_count} новых.")
                return redirect('upload_teacher_assignments')

            except Exception as e:
                error_msg = f"Ошибка при обработке файла: {e}"
                messages.error(request, error_msg)

                # Удаляем файл и запись из БД
                try:
                    if 'assignment_file' in locals():
                        assignment_file.delete()
                    if os.path.exists(full_path):
                        os.remove(full_path)
                except Exception as cleanup_err:
                    messages.warning(request, f"Не удалось удалить файл: {cleanup_err}")

                # Вместо редиректа возвращаем render, чтобы ошибки не терялись
                assignment_files = TeacherAssignmentFile.objects.select_related('uploaded_by').order_by('-upload_date')
                data = [
                    {
                        'assignment': af,
                        'teacher_count': TeachingAssignment.objects.filter(assignment_file=af).count()
                    } for af in assignment_files
                ]
                return render(request, 'upload_assignments.html', {
                    'form': form,
                    'assignments_data': data
                })

    # GET или первый вызов
    assignment_files = TeacherAssignmentFile.objects.select_related('uploaded_by').order_by('-upload_date')
    data = [
        {
            'assignment': af,
            'teacher_count': TeachingAssignment.objects.filter(assignment_file=af).count()
        } for af in assignment_files
    ]
    context['assignments_data'] = data
    return render(request, 'upload_assignments.html', context)




def match_user_by_initials(fio_initials, users):
    import re
    match = re.match(r'^([А-ЯЁA-Z][а-яёa-z]+)\s+([А-ЯЁA-Z])\.([А-ЯЁA-Z])\.$', fio_initials)
    if not match:
        return None

    last_name, first_init, middle_init = match.groups()

    for profile in users:
        if (
            profile.last_name.lower() == last_name.lower()
            and profile.first_name and profile.first_name[0].upper() == first_init.upper()
            and profile.middle_name and profile.middle_name[0].upper() == middle_init.upper()
        ):
            return profile.user
    return None


def parse_teacher_assignments(user, file_path, assignment_file):
    import os
    print(">> Путь к файлу:", file_path)
    print(">> Файл существует:", os.path.exists(file_path))

    if not os.path.exists(file_path):
        raise ValueError(f"Файл не найден: {file_path}")

    profile = ColorProfile.objects.get(user=user)

    def hex_to_rgb(hex_color):
        return "FF" + hex_color.lstrip("#").upper()

    required_colors = [
        profile.ta_teacher_color,
        profile.ta_subject_color,
        profile.ta_group_color
    ]
    if not all(required_colors):
        raise ValueError("Не заданы все необходимые цвета в настройках завуча.")

    color_map = {
        hex_to_rgb(profile.ta_teacher_color): 'teacher',
        hex_to_rgb(profile.ta_subject_color): 'subject',
        hex_to_rgb(profile.ta_group_color): 'group',
    }
    recognized_colors = set(color_map.keys())

    try:
        wb = load_workbook(file_path)
    except Exception as e:
        raise ValueError(f"Ошибка при чтении Excel-файла: {e}")

    assignments = []
    unmatched_teachers = set()
    users = list(UserProfile.objects.select_related('user'))

    with transaction.atomic():
        for sheet in wb.worksheets:
            print(f">>> Обрабатывается лист: {sheet.title}")
            current_teacher = None
            current_subject = None
            current_group = None

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    rgb = cell.fill.start_color.rgb
                    if rgb is None or rgb not in recognized_colors:
                        continue
                    cell_type = color_map[rgb]
                    val = str(cell.value).strip()

                    if cell_type == 'teacher':
                        current_teacher = val
                    elif cell_type == 'subject':
                        current_subject = val
                    elif cell_type == 'group':
                        current_group = val

                if current_teacher and current_subject and current_group:
                    teacher_user = match_user_by_initials(current_teacher, users)
                    if not teacher_user:
                        print(f"[X] Преподаватель не найден: {current_teacher}")
                        unmatched_teachers.add(current_teacher)
                        continue

                    subject_obj, _ = Subject.objects.get_or_create(name=current_subject)
                    TeachingAssignment.objects.update_or_create(
                        subject=subject_obj,
                        group=current_group,
                        teacher=teacher_user,  # <-- Включаем в ключ уникальности
                        defaults={'assignment_file': assignment_file}
                    )
                    assignments.append((teacher_user.username, current_subject, current_group))
                    current_teacher = current_subject = current_group = None

    print(f">>> Завершено. Назначений: {len(assignments)}")
    print(f">>> Преподаватели без совпадения: {unmatched_teachers}")

    if unmatched_teachers:
        raise ValueError(
            f"Файл не загружен: не найдены преподаватели: {', '.join(unmatched_teachers)}"
        )

    if not assignments:
        raise ValueError("Файл не содержит ни одной записи для сохранения.")

    return len(assignments)


def semester_index(academic_year, semester):
    start_year = int(academic_year.split('-')[0])
    return (start_year - 2000) * 2 + int(semester)

def generate_excel_report(group_name, from_year, to_year):
    def semester_index(academic_year, semester):
        start_year = int(academic_year.split('-')[0])
        return (start_year - 2000) * 2 + int(semester)

    start_index = semester_index(from_year, '1')
    end_index = semester_index(to_year, '2')

    vedomosti = Vedomost.objects.filter(group_name=group_name)
    vedomosti = [
        v for v in vedomosti
        if start_index <= semester_index(v.academic_year, v.semester) <= end_index
    ]

    if not vedomosti:
        raise ValueError("Нет данных за выбранный период.")

    grades = Grade.objects.filter(vedomost__in=vedomosti).select_related('student')

    if not grades.exists():
        raise ValueError("Нет оценок за выбранный период.")

    student_stats = {}
    for grade in grades:
        student = grade.student
        if student.full_name not in student_stats:
            student_stats[student.full_name] = {
                'student': student,
                'grades': []
            }
        student_stats[student.full_name]['grades'].append(grade.value)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Отчет {group_name}"

    bold_font = Font(bold=True)
    red_font = Font(color="FF0000")
    section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    headers = ['ФИО', 'Средний балл', '5', '4', '3', '2', 'Характеристика']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="center")

    for stats in student_stats.values():
        grades_list = stats['grades']
        numeric = [int(g) for g in grades_list if g.isdigit()]
        avg = round(sum(numeric) / len(numeric), 2) if numeric else 0

        count_5 = numeric.count(5)
        count_4 = numeric.count(4)
        count_3 = numeric.count(3)
        count_2 = numeric.count(2)
        fail_count = count_2

        if fail_count > 3:
            status = "Задолжник"
        elif avg >= 4.75:
            status = "Отличник"
        elif avg >= 4.0:
            status = "Хорошист"
        elif avg >= 3.0:
            status = "Удовлетворительно"
        else:
            status = "Задолжник"

        row = [
            stats['student'].full_name,
            avg,
            count_5,
            count_4,
            count_3,
            count_2,
            status
        ]
        ws.append(row)

        if fail_count > 3:
            for cell in ws[ws.max_row]:
                cell.font = red_font

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                value_length = len(str(cell.value))
                if value_length > max_length:
                    max_length = value_length
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=group_report_{group_name}.xlsx'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.read())
    return response

def generate_student_report(student, from_year, to_year):
    def semester_index(academic_year, semester):
        start_year = int(academic_year.split('-')[0])
        return (start_year - 2000) * 2 + int(semester)

    start_index = semester_index(from_year, '1')
    end_index = semester_index(to_year, '2')

    vedomosti = Vedomost.objects.filter(group_name=student.group)
    vedomosti = [
        v for v in vedomosti
        if start_index <= semester_index(v.academic_year, v.semester) <= end_index
    ]

    if not vedomosti:
        raise ValueError("Нет ведомостей за выбранный период.")

    grades = Grade.objects.filter(student=student, vedomost__in=vedomosti).select_related('subject')

    if not grades.exists():
        raise ValueError("У студента нет оценок за выбранный период.")

    numeric = [int(g.value) for g in grades if g.value.isdigit()]
    avg = round(sum(numeric) / len(numeric), 2) if numeric else 0

    count_5 = numeric.count(5)
    count_4 = numeric.count(4)
    count_3 = numeric.count(3)
    count_2 = numeric.count(2)
    fail_count = count_2

    if fail_count > 3:
        status = "Задолжник"
    elif avg >= 4.75:
        status = "Отличник"
    elif avg >= 4.0:
        status = "Хорошист"
    elif avg >= 3.0:
        status = "Удовлетворительно"
    else:
        status = "Задолжник"

    bad_subjects = grades.filter(value='2').values_list('subject__name', flat=True).distinct()

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по студенту"

    bold_font = Font(bold=True)
    red_font = Font(color="FF0000")
    section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    center_align = Alignment(horizontal="left")

    ws.append(["ФИО", student.full_name])
    ws.append(["Группа", student.group])
    ws.append(["Период", f"{from_year} — {to_year}"])
    ws.append([])

    ws.append(["Статистика"])
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=2)
    ws["A5"].font = bold_font
    ws["A5"].fill = section_fill

    data = [
        ["Средний балл", avg],
        ["Пятёрок", count_5],
        ["Четвёрок", count_4],
        ["Троек", count_3],
        ["Двоек", count_2],
        ["Характеристика", status]
    ]

    for row in data:
        ws.append(row)

    ws.append([])

    if bad_subjects:
        ws.append(["Предметы с оценкой '2'"])
        ws["A{}".format(ws.max_row)].font = bold_font
        ws["A{}".format(ws.max_row)].fill = section_fill
        for subject in bad_subjects:
            cell = ws.cell(row=ws.max_row + 1, column=1, value=subject)
            cell.font = red_font

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                value_length = len(str(cell.value))
                if value_length > max_length:
                    max_length = value_length
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=student_report_{student.full_name}.xlsx'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.read())
    return response

def generate_teacher_report(teacher, from_year, to_year):
    def semester_index(academic_year, semester):
        start_year = int(academic_year.split('-')[0])
        return (start_year - 2000) * 2 + int(semester)

    start_index = semester_index(from_year, '1')
    end_index = semester_index(to_year, '2')

    assignments = TeachingAssignment.objects.filter(teacher=teacher).select_related('subject')
    if not assignments.exists():
        raise ValueError("У преподавателя нет назначений.")

    filters = Q()
    for ta in assignments:
        filters |= Q(subject=ta.subject, vedomost__group_name=ta.group)

    grades = Grade.objects.filter(filters).select_related('vedomost', 'student', 'subject')

    filtered_grades = []
    for g in grades:
        ved = g.vedomost
        if ved.academic_year and ved.semester:
            index = semester_index(ved.academic_year, ved.semester)
            if start_index <= index <= end_index:
                filtered_grades.append(g)

    if not filtered_grades:
        raise ValueError("Нет оценок за выбранный период.")

    stat_table = defaultdict(lambda: {'5': 0, '4': 0, '3': 0, '2': 0})
    fail_students = []

    for g in filtered_grades:
        group = g.vedomost.group_name
        subject = g.subject.name
        value = g.value.strip()

        if value in ('5', '4', '3', '2'):
            stat_table[(group, subject)][value] += 1
            if value == '2':
                fail_students.append((g.student.full_name, group, subject))

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт по преподавателю"

    bold_font = Font(bold=True)
    red_font = Font(color="FF0000")
    section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")

    ws.append(["ФИО преподавателя", f"{teacher.userprofile.last_name} {teacher.userprofile.first_name}"])
    ws.append(["Период", f"{from_year} — {to_year}"])
    ws.append([])

    ws.append(["Группа", "Предмет", "5", "4", "3", "2"])
    for cell in ws[ws.max_row]:
        cell.font = bold_font
        cell.fill = section_fill

    for (group, subject), stats in stat_table.items():
        ws.append([group, subject, stats['5'], stats['4'], stats['3'], stats['2']])

    ws.append([])

    if fail_students:
        ws.append(["Студенты с оценкой '2'"])
        ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=3)
        ws[ws.max_row][0].font = bold_font
        ws[ws.max_row][0].fill = section_fill

        ws.append(["ФИО", "Группа", "Предмет"])
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        for full_name, group, subject in fail_students:
            ws.append([full_name, group, subject])
            for cell in ws[ws.max_row]:
                cell.font = red_font

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_len + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=teacher_report_{teacher.username}.xlsx'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.read())
    return response

@login_required
def generate_report_view(request):
    if request.user.userprofile.role != 'deputy':
        messages.error(request, "Доступ запрещён.")
        return redirect('home')

    vedomosti = Vedomost.objects.exclude(academic_year__isnull=True)
    groups = sorted(set(vedomosti.values_list('group_name', flat=True)))
    years = sorted(set(vedomosti.values_list('academic_year', flat=True)))
    students = Student.objects.order_by('full_name')
    teachers = User.objects.filter(
        id__in=TeachingAssignment.objects.values_list('teacher_id', flat=True)
    ).distinct().select_related('userprofile')

    form = ReportGenerationForm(
        request.POST or None,
        groups=groups,
        students=students,
        years=years,
        teachers=teachers,
    )

    if request.method == 'POST' and form.is_valid():
        report_type = form.cleaned_data['report_type']
        from_year = form.cleaned_data['from_year']
        to_year = form.cleaned_data['to_year']

        try:
            if report_type == 'group':
                group = form.cleaned_data['group']
                return generate_excel_report(group, from_year, to_year)
            elif report_type == 'student':
                student_id = form.cleaned_data['student']
                student = Student.objects.get(id=student_id)
                return generate_student_report(student, from_year, to_year)
            elif report_type == 'teacher':
                teacher_id = form.cleaned_data['teacher']
                teacher = User.objects.get(id=teacher_id)
                return generate_teacher_report(teacher, from_year, to_year)
        except ValueError as e:
            messages.error(request, str(e))

    return render(request, 'generate_report.html', {'form': form})

@login_required
def ajax_get_years_by_entity(request):
    teacher_id = request.GET.get('teacher')
    group = request.GET.get('group')
    student_id = request.GET.get('student')

    if teacher_id:
        tas = TeachingAssignment.objects.filter(teacher__id=teacher_id).select_related('subject')
        if not tas.exists():
            return JsonResponse({'years': []})

        valid_pairs = {(ta.group, ta.subject.name) for ta in tas}
        grades = Grade.objects.select_related('vedomost', 'subject')
        valid_ved_ids = set()

        for g in grades:
            if (g.vedomost.group_name, g.subject.name) in valid_pairs:
                valid_ved_ids.add(g.vedomost.id)

        vedomosti = Vedomost.objects.filter(id__in=valid_ved_ids)
        years = sorted(set(v.academic_year for v in vedomosti if v.academic_year))
        return JsonResponse({'years': years})

    elif group:
        vedomosti = Vedomost.objects.filter(group_name=group)
    elif student_id:
        student = Student.objects.filter(id=student_id).first()
        vedomosti = Vedomost.objects.filter(group_name=student.group) if student else []
    else:
        return JsonResponse({'error': 'Не передан идентификатор'}, status=400)

    years = sorted(set(v.academic_year for v in vedomosti if v.academic_year))
    return JsonResponse({'years': years})

@login_required
def users_list(request):
    if request.user.userprofile.role != 'deputy':
        messages.error(request, "Доступ запрещён.")
        return redirect('home')

    users = UserProfile.objects.select_related('user').all()
    return render(request, 'users_list.html', {'users': users})

@login_required
def edit_user(request, user_id):
    if request.user.userprofile.role != 'deputy':
        messages.error(request, "Доступ запрещён.")
        return redirect('home')

    user = get_object_or_404(User, id=user_id)
    profile = user.userprofile

    if request.method == 'POST':
        form = EditUserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            profile.first_name = form.cleaned_data['first_name']
            profile.last_name = form.cleaned_data['last_name']
            profile.middle_name = form.cleaned_data['middle_name']
            profile.role = form.cleaned_data['role']
            profile.save()
            messages.success(request, "Пользователь обновлён.")
            return redirect('users_list')
    else:
        form = EditUserForm(instance=user, initial={
            'first_name': profile.first_name,
            'last_name': profile.last_name,
            'middle_name': profile.middle_name,
            'role': profile.role
        })

    return render(request, 'edit_user.html', {'form': form, 'user_obj': user})

@login_required
def delete_user(request, user_id):
    if request.user.userprofile.role != 'deputy':
        messages.error(request, "Доступ запрещён.")
        return redirect('home')

    user = get_object_or_404(User, id=user_id)

    if request.method == 'POST':
        user.delete()
        messages.success(request, "Пользователь удалён.")
        return redirect('users_list')

    return render(request, 'confirm_delete.html', {'user_obj': user})

@login_required
def delete_teacher_assignment(request, pk):
    if request.user.userprofile.role != 'deputy':
        messages.error(request, "Доступ запрещён.")
        return redirect('home')

    try:
        file_obj = TeacherAssignmentFile.objects.get(pk=pk)
        path = os.path.join(settings.MEDIA_ROOT, file_obj.file.name)

        # Удаляем все TeachingAssignment, связанные с этим файлом
        TeachingAssignment.objects.filter(assignment_file=file_obj).delete()

        # Удаляем файл с диска
        if os.path.exists(path):
            os.remove(path)

        # Удаляем запись о файле
        file_obj.delete()

        messages.success(request, "Файл и все связанные назначения успешно удалены.")
    except TeacherAssignmentFile.DoesNotExist:
        messages.error(request, "Файл не найден.")
    return redirect('upload_teacher_assignments')

